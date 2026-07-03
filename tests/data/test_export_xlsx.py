from openpyxl import load_workbook

from rads_explorer.data.export.xlsx import XLSXExporter


def test_xlsx_export(report_service, tmp_path):
    report = report_service.certificates_inventory_report()

    output = tmp_path / "report.xlsx"

    XLSXExporter().export(report, output)

    assert output.exists()
    assert output.stat().st_size > 0

    wb = load_workbook(output)
    ws = wb.active

    assert ws.title == report.name

    headers = [cell.value for cell in ws[1]]
    assert "serial_number" in headers
    assert "guid" in headers
