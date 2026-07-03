from pathlib import Path

from openpyxl import load_workbook

from rads_explorer.data.export.xlsx import XLSXExporter
from rads_explorer.data.reports import Report
from rads_explorer.data.reports_models import CertificateReportRow


def test_xlsx_export(tmp_path: Path):
    row = CertificateReportRow(
        ogrn="123",
        organization_name="Org",
        guid="gid",
        surname="Ivanov",
        given_name="Ivan",
        organizational_unit_name="IT",
        title="Engineer",
        common_name="CN",
        serial_number="ABC",
        snils="123",
        status="VALID",
        certificate_template="1.2.3",
        revoked_when=None,
        not_before="2024-01-01",
        not_after="2025-01-01",
    )

    report = Report(
        name="test_report",
        generated_at=None,
        data=[row],
    )

    output_file = tmp_path / "report.xlsx"

    XLSXExporter().export(report, output_file)

    wb = load_workbook(output_file)
    ws = wb.active

    assert ws.title == "test_report"

    rows = list(ws.rows)

    assert [c.value for c in rows[0]] == list(row.__dataclass_fields__.keys())


def test_xlsx_exporter_generates_valid_report(tmp_path: Path):
    row = [
        CertificateReportRow(
            ogrn="123",
            organization_name="TestOrg",
            guid="gid-1",
            surname="Ivanov",
            given_name="Ivan Ivanovich",
            organizational_unit_name="IT",
            title="Engineer",
            common_name="Ivanov I.I.",
            serial_number="ABC123",
            snils="999",
            status="VALID",
            certificate_template="1.2.3",
            revoked_when=None,
            not_before="2024-01-01T00:00:00Z",
            not_after="2025-01-01T00:00:00Z",
        )
    ]

    report = Report(
        name="test_report",
        generated_at="2024-01-01T00:00:00Z",
        data=row,
    )

    output_file = tmp_path / "report.xlsx"

    XLSXExporter().export(report, output_file)

    assert output_file.exists()

    wb = load_workbook(output_file)
    ws = wb.active

    # header row
    headers = [c.value for c in ws[1]]
    assert "ogrn" in headers
    assert "organization_name" in headers
    assert "serial_number" in headers

    # data row exists
    row = [c.value for c in ws[2]]
    assert "TestOrg" in row
    assert "ABC123" in row
